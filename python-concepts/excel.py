import json
import re
import pandas as pd
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def parse_html_content(html_content):
    """
    Parse HTML content to extract student information and subject details
    """
    student_data = {
        'reg_no': '',
        'name': '',
        'course': '',
        'credits_earned': '',
        'cgpa': '',
        'sgpa': '',
        'non_academic_credits': '',
        'subjects': []
    }
    
    # Extract student basic information
    try:
        # Use regex to find student info more reliably
        reg_match = re.search(r'<td ALIGN="middle">([A-Z0-9]+)</td>', html_content)
        name_match = re.search(r'<td ALIGN="middle">[A-Z0-9]+</td>\s*<td ALIGN="middle">([^<]+)</td>', html_content)
        course_match = re.search(r'<td ALIGN="middle">([^<]*M\.Sc[^<]*)</td>', html_content)
        
        if reg_match:
            student_data['reg_no'] = reg_match.group(1).strip()
        if name_match:
            student_data['name'] = name_match.group(1).strip()
        if course_match:
            student_data['course'] = course_match.group(1).replace('&nbsp;', ' ').strip()
            
    except Exception as e:
        print(f"Error extracting student info: {e}")
    
    # Extract CGPA, SGPA, Credits
    try:
        blue_values = re.findall(r'<FONT COLOR="blue"[^>]*><b>([^<]+)</b></FONT>', html_content, re.IGNORECASE)
        
        if len(blue_values) >= 3:
            student_data['credits_earned'] = blue_values[0]
            student_data['cgpa'] = blue_values[1]
            student_data['sgpa'] = blue_values[2]
            if len(blue_values) >= 4:
                student_data['non_academic_credits'] = blue_values[3]
    except Exception as e:
        print(f"Error extracting grades info: {e}")
    
    # Extract subject information
    try:
        subject_pattern = r'<td align="middle"><font color="black">\s*(\d+)\s*</font></td>\s*<td align="middle"><font color="black">\s*([^<]+)\s*</font></td>\s*<td>\s*<font color="black">\s*([^<]+)\s*</font></td>\s*<td align="middle"><font color="BLACK">([^<]+)</font></td>\s*<td align="middle"><font color="black">([^<]+)</font></td>'
        
        subject_matches = re.findall(subject_pattern, html_content, re.IGNORECASE | re.DOTALL)
        
        for match in subject_matches:
            subject = {
                'semester': match[0].strip(),
                'code': match[1].strip(),
                'subject': match[2].strip(),
                'grade': match[3].strip(),
                'credit': match[4].strip()
            }
            student_data['subjects'].append(subject)
            
    except Exception as e:
        print(f"Error extracting subjects: {e}")
    
    return student_data

def create_horizontal_excel(students_data, output_filename='exam_results_horizontal.xlsx'):
    """
    Create an Excel file with horizontal format (one row per student)
    """
    # Collect all unique subjects across all students
    all_subjects = {}
    subject_order = []
    
    for student in students_data:
        for subject in student['subjects']:
            subject_key = f"{subject['code']}-{subject['subject']}"
            if subject_key not in all_subjects:
                all_subjects[subject_key] = subject_key
                subject_order.append(subject_key)
    
    # Create the data structure for horizontal format
    excel_rows = []
    
    for idx, student in enumerate(students_data, 1):
        # Start with basic info
        row_data = {
            'S.No.': idx,
            'Reg.No.': student['reg_no'],
            'Name of the student': student['name']
        }
        
        # Add subject grades
        student_subjects = {f"{s['code']}-{s['subject']}": s['grade'] for s in student['subjects']}
        
        for subject_key in subject_order:
            row_data[subject_key] = student_subjects.get(subject_key, '')
        
        # Add summary columns
        current_year_arrear = sum(1 for s in student['subjects'] if s['grade'] in ['P', 'F', 'Ab'])
        previous_year_arrear = 0  # This would need to be calculated based on your specific logic
        total_arrears = current_year_arrear + previous_year_arrear
        
        row_data.update({
            'Current year arrear': current_year_arrear,
            'Previous year arrear': previous_year_arrear,
            'Total': total_arrears,
            'non academic credit': student['non_academic_credits'],
            'CGPA': student['cgpa']
        })
        
        excel_rows.append(row_data)
    
    # Create DataFrame
    df = pd.DataFrame(excel_rows)
    
    # Create Excel workbook with formatting
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Results"
    
    # Define styles
    header_font = Font(name='Arial', size=10, bold=True, color='000000')
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    data_font = Font(name='Arial', size=10)
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write data to worksheet
    for r_idx, (_, row) in enumerate(df.iterrows(), 1):
        for c_idx, (col_name, value) in enumerate(row.items(), 1):
            cell = ws.cell(row=r_idx+1, column=c_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write headers
    for c_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Set minimum width for subject columns
        if any(keyword in str(column[0].value) for keyword in ['20CA', '17VE', '22MA', '23CA']):
            adjusted_width = max(8, min(max_length + 1, 12))
        else:
            adjusted_width = max(10, min(max_length + 2, 30))
        
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the file
    wb.save(output_filename)
    print(f"Horizontal Excel file saved as: {output_filename}")

def main(json_data=None, input_file=None):
    """
    Main function to process JSON data and create horizontal Excel output
    """
    try:
        if json_data:
            if isinstance(json_data, str):
                data = json.loads(json_data)
            else:
                data = json_data
        elif input_file:
            with open(input_file, 'r', encoding='utf-8') as file:
                content = file.read()
            data = json.loads(content)
        else:
            # Use embedded sample data
            json_string = '''[
{
  "reg_no": "PRK24AD1007",    
  "raw_html": "<html>\\r\\n<head>\\r\\n\\r\\n<title>KITS Exam - Apr/May 2025 Results</title></head>\\r\\n<body  leftmargin=0 topmargin=0 marginheight=\\"0\\" marginwidth=\\"0\\" style=\\"background-attachment: fixed\\" bgcolor=#547DEB background=\\"t2.jpg\\">\\r\\n\\r\\n\\r\\n\\r\\n<center>\\r\\n<img src=\\"http://www.karunya.edu/sites/default/files/logo.png.pagespeed.ce.rQZd6s4jcx.png\\" alt=\\"Porto\\"></center>\\r\\n<TABLE WIDTH=\\"87%\\" BORDER=0 CELLSPACING=1 CELLPADDING=1>\\r\\n\\t<TR>\\r\\n\\r\\n\\t\\t\\r\\n\\t\\t<TD width=\\"1044\\">\\r\\n\\t\\t<h2><font color=\\"#0000FF\\" face=\\"Book Antiqua\\">Exam Results - Apr/May \\r\\n\\t\\t2025 (</font><font face=\\"Book Antiqua\\" style=\\"font-size: 15pt\\" color=\\"#008000\\">All</font><font color=\\"#008000\\" face=\\"Book Antiqua\\"><span style=\\"font-size: 15pt\\"> \\r\\n\\t\\tStudents</span></font><font color=\\"#0000FF\\" face=\\"Book Antiqua\\\">) </font>\\r\\n\\t\\t\\r\\n\\t\\t</h2>  \\r\\n\\t\\t</TD>\\r\\n\\t\\t<TD>\\r\\n\\t\\t<p align=\\"center\\"><b><a href=\\"logout_new5.asp\\">LOGOUT</a></b></TD>\\t\\t\\r\\n\\t</TR>\\r\\n</TABLE>\\r\\n\\r\\n\\r\\n\\r\\n\\r\\n\\r\\n\\r\\n<TABLE border=3 borderColorDark=#c9beb6 borderColorLight=#000000 cellPadding=1 \\r\\ncellSpacing=1 width=\\"70%\\" align=center>\\r\\n\\r\\n\\t<tr>\\r\\n\\t\\t<td ALIGN=\\"middle\\"><b>REG. NO.</b></td>\\r\\n\\t\\t<td ALIGN=\\"middle\\"><b>NAME</b></td>\\r\\n\\t\\t<td ALIGN=\\"middle\\"><b>COURSE</b></td>\\r\\n\\r\\n\\t</tr>\\r\\n\\r\\n\\t<tr>\\r\\n\\t\\t<td ALIGN=\\"middle\\">PRK24AD1007</td>\\r\\n\\t\\t<td ALIGN=\\"middle\\">SANJITH KRISHNA B</td>\\r\\n\\t\\t<td ALIGN=\\"middle\\">M.Sc.&nbsp;&nbsp;Artificial Intelligence and Data Science</td>\\r\\n\\t\\r\\n\\t</tr>\\r\\n</TABLE>\\r\\n<CENTER>\\r\\n<TABLE>\\r\\n\\t<TR>\\r\\n\\t\\t<TD>\\r\\n\\t\\t\\t<table BORDER=1 ALIGN=center>\\r\\n\\t\\t\\t\\t<tr>\\r\\n\\t\\t\\t\\t\\t<td WIDTH=150 ALIGN=middle>Credits Earned</td> \\r\\n\\t\\t\\t\\t\\t<td WIDTH=150 ALIGN=middle>CGPA</td> \\r\\n\\t\\t\\t\\t</tr>\\r\\n\\t\\t\\t\\t<tr>\\r\\n\\t\\t\\t\\t\\t\\r\\n\\t\\t\\t\\t\\t<TD ALIGN=middle><FONT COLOR=\\"blue\\" ><b>47</b></FONT></TD>\\r\\n\\t\\t\\t\\t\\t\\r\\n\\t\\t\\t\\t\\t<TD ALIGN=middle><FONT COLOR=\\"blue\\"><b>6.74</b></FONT></TD>\\r\\n\\t\\t\\t\\t\\t\\r\\n\\r\\n\\t\\t\\t\\t</tr>\\r\\n\\t\\t\\t</table>\\r\\n\\t\\t<TD WIDTH=10></TD>\\r\\n\\t\\t<TD>\\r\\n\\t\\t\\t<table BORDER=1 ALIGN=center>\\r\\n\\t\\t\\t\\t<TR>\\r\\n\\t\\t\\t\\t\\t<td WIDTH=150 ALIGN=middle>SGPA</td>\\r\\n\\t\\t\\t\\t</TR>\\r\\n\\t\\t\\t\\t<tr>\\r\\n\\t\\t\\t\\t\\t\\r\\n\\t\\t\\t\\t\\t<TD ALIGN=middle><FONT COLOR=\\"blue\\"><b>6.54</b></FONT></TD>\\r\\n\\t\\t\\t\\t\\t\\r\\n\\t\\t\\t\\t\\t\\r\\n\\t\\t\\t\\t\\t\\r\\n\\t\\t\\t\\t\\t\\r\\n\\t\\t\\t\\t</tr>\\r\\n\\t\\t\\t</table>\\r\\n\\t\\t</TD>\\r\\n\\t\\t<TD WIDTH=50></TD>\\r\\n\\r\\n\\t\\t<td>\\r\\n\\t\\t<table BORDER=1 ALIGN=center>\\r\\n\\t\\t\\t\\t<TR>\\r\\n\\t\\t\\t\\t\\t<td WIDTH=150 ALIGN=middle>Non Academic Credits Earned</td>\\r\\n\\t\\t\\t\\t</TR>\\r\\n\\t\\t\\t\\t<tr>\\r\\n\\t\\t\\t\\t\\t<TD ALIGN=middle><FONT COLOR=\\"blue\\"><b>1</b></FONT></TD>\\r\\n\\t\\t\\t\\t\\t\\r\\n\\t\\t\\t\\t\\t\\r\\n\\t\\t\\t\\t\\t\\r\\n\\t\\t\\t\\t</tr>\\r\\n\\t\\t\\t</table>\\r\\n\\r\\n\\t\\t\\r\\n\\t\\t\\r\\n\\t\\t</td>\\r\\n\\t\\t\\r\\n\\t\\t\\r\\n\\t\\t\\r\\n\\t\\t\\r\\n\\t</TR>\\r\\n</TABLE>\\r\\n\\t\\r\\n\\t\\r\\n\\t\\r\\n</CENTER>\\r\\n<small>\\r\\n\\r\\n\\r\\n<!--\\r\\n<CENTER><FONT size=3><STRONG>\\r\\n<FONT color=green >Congratulations ! &nbsp;&nbsp;&nbsp; You have cleared all papers !</FONT>  </FONT>\\r\\n</CENTER></STRONG>\\r\\n-->\\r\\n\\r\\n\\r\\n<center>\\r\\n\\r\\n<TABLE border=3 borderColorDark=#c9beb6 borderColorLight=#000000 cellPadding=1 \\r\\ncellSpacing=1 width=\\"65%\\">\\r\\n\\r\\n\\r\\n  \\r\\n<tbody>\\r\\n\\r\\n\\r\\n\\r\\n\\r\\n\\r\\n<tr ALIGN=\\"middle\\"><td><b>SEM</b></td><td><b>CODE</b></td><td><b>SUBJECT</b></td>\\r\\n<td><b>GRADE</b></td><td><b><b>CREDIT</b></b></td></tr></B>\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">  2</font></td>\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">  23CA3034</font></td>\\r\\n\\r\\n<td> <font color=\\"black\\"> Data Warehousing and Data Mining Techniques</font></td>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"BLACK\\">P</font></td>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">3</font></td>\\r\\n\\r\\n\\r\\n\\r\\n<td ALIGN=\\"middle\\"><font color=\\"black\\" \\r\\n      style=\\"COLOR: fuchsia\\">&nbsp; </font></td>\\r\\n\\r\\n\\r\\n</tr>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">  2</font></td>\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">  23CA3036</font></td>\\r\\n\\r\\n<td> <font color=\\"black\\"> Data Analytics and Visualization</font></td>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"BLACK\\">B+</font></td>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">3</font></td>\\r\\n\\r\\n\\r\\n\\r\\n<td ALIGN=\\"middle\\"><font color=\\"black\\" \\r\\n      style=\\"COLOR: fuchsia\\">&nbsp; </font></td>\\r\\n\\r\\n\\r\\n</tr>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">  2</font></td>\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">  23CA3038</font></td>\\r\\n\\r\\n<td> <font color=\\"black\\"> Advanced Database Technologies</font></td>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"BLACK\\">B</font></td>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">3</font></td>\\r\\n\\r\\n\\r\\n\\r\\n<td ALIGN=\\"middle\\"><font color=\\"black\\" \\r\\n      style=\\"COLOR: fuchsia\\">&nbsp; </font></td>\\r\\n\\r\\n\\r\\n</tr>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">  2</font></td>\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">  23CA3039</font></td>\\r\\n\\r\\n<td> <font color=\\"black\\"> Human Centered Computing</font></td>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"BLACK\\">B+</font></td>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">3</font></td>\\r\\n\\r\\n\\r\\n\\r\\n<td ALIGN=\\"middle\\"><font color=\\"black\\" \\r\\n      style=\\"COLOR: fuchsia\\">&nbsp; </font></td>\\r\\n\\r\\n\\r\\n</tr>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">  2</font></td>\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">  23CA3040</font></td>\\r\\n\\r\\n<td> <font color=\\"black\\"> Modelling Techniques in Predictive Analytics</font></td>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"BLACK\\">B+</font></td>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">3</font></td>\\r\\n\\r\\n\\r\\n\\r\\n<td ALIGN=\\"middle\\"><font color=\\"black\\" \\r\\n      style=\\"COLOR: fuchsia\\">&nbsp; </font></td>\\r\\n\\r\\n\\r\\n</tr>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">  2</font></td>\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">  23CA3042</font></td>\\r\\n\\r\\n<td> <font color=\\"black\\"> Internet of Things and Blockchain Technology</font></td>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"BLACK\\">P</font></td>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">3</font></td>\\r\\n\\r\\n\\r\\n\\r\\n<td ALIGN=\\"middle\\"><font color=\\"black\\" \\r\\n      style=\\"COLOR: fuchsia\\">&nbsp; </font></td>\\r\\n\\r\\n\\r\\n</tr>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">  2</font></td>\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">  22MA3001</font></td>\\r\\n\\r\\n<td> <font color=\\"black\\"> Logical Reasoning and Soft Skills</font></td>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"BLACK\\">Completed</font></td>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">0</font></td>\\r\\n\\r\\n\\r\\n\\r\\n<td ALIGN=\\"middle\\"><font color=\\"black\\" \\r\\n      style=\\"COLOR: fuchsia\\">&nbsp; </font></td>\\r\\n\\r\\n\\r\\n</tr>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">  2</font></td>\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">  23CA3035</font></td>\\r\\n\\r\\n<td> <font color=\\"black\\"> Machine Learning Lab</font></td>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"BLACK\\">B+</font></td>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">2</font></td>\\r\\n\\r\\n\\r\\n\\r\\n<td ALIGN=\\"middle\\"><font color=\\"black\\" \\r\\n      style=\\"COLOR: fuchsia\\">&nbsp; </font></td>\\r\\n\\r\\n\\r\\n</tr>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">  2</font></td>\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">  23CA3037</font></td>\\r\\n\\r\\n<td> <font color=\\"black\\"> Data Analytics and Visualization Lab</font></td>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"BLACK\\">O</font></td>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">2</font></td>\\r\\n\\r\\n\\r\\n\\r\\n<td ALIGN=\\"middle\\"><font color=\\"black\\" \\r\\n      style=\\"COLOR: fuchsia\\">&nbsp; </font></td>\\r\\n\\r\\n\\r\\n</tr>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">  2</font></td>\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">  23CA3041</font></td>\\r\\n\\r\\n<td> <font color=\\"black\\"> Predictive Analytics Lab</font></td>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"BLACK\\">A+</font></td>\\r\\n\\r\\n\\r\\n<td align=\\"middle\\"><font color=\\"black\\">2</font></td>\\r\\n\\r\\n\\r\\n\\r\\n<td ALIGN=\\"middle\\"><font color=\\"black\\" \\r\\n      style=\\"COLOR: fuchsia\\">&nbsp; </font></td>\\r\\n\\r\\n\\r\\n</tr>"
}
]'''
            data = json.loads(json_string)
        
        print(f"Found {len(data)} student records")
        
        # Parse each student's HTML content
        students_data = []
        for record in data:
            if 'reg_no' in record and 'raw_html' in record:
                student_info = parse_html_content(record['raw_html'])
                # Use reg_no from JSON if not found in HTML
                if not student_info['reg_no']:
                    student_info['reg_no'] = record['reg_no']
                students_data.append(student_info)
                
                print(f"Processed student: {student_info['name']} ({student_info['reg_no']}) - {len(student_info['subjects'])} subjects")
        
        # Create horizontal Excel file
        if students_data:
            create_horizontal_excel(students_data, 'student_results_horizontal.xlsx')
            print("Horizontal format Excel file created successfully!")
        else:
            print("No valid student data found!")
            
    except FileNotFoundError:
        print(f"Error: Input file not found!")
    except json.JSONDecodeError as e:
        print(f"Error parsing JSON: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    # Required libraries installation instructions
    print("Required libraries: pandas, openpyxl, beautifulsoup4")
    print("Install with: pip install pandas openpyxl beautifulsoup4")
    print("-" * 70)
    
    # Run with embedded sample data
    main()

# Example of how to use with your own data:
def process_custom_data():
    """
    Example: Replace this with your actual JSON data
    """
    your_json_data = '''[
{
  "reg_no": "YOUR_REG_NO",    
  "raw_html": "YOUR_HTML_CONTENT_HERE"
}
]'''
    
    main(json_data=your_json_data)

# If you have multiple students, just add more objects to the array:
def process_multiple_students():
    """
    Example for multiple students
    """
    multiple_students_json = '''[
{
  "reg_no": "PRK24AD1007",    
  "raw_html": "<html>\r\n<head>\r\n\r\n<title>KITS Exam - Apr/May 2025 Results</title></head>\r\n<body  leftmargin=0 topmargin=0 marginheight=\"0\" marginwidth=\"0\" style=\"background-attachment: fixed\" bgcolor=#547DEB background=\"t2.jpg\">\r\n\r\n\r\n\r\n<center>\r\n<img src=\"http://www.karunya.edu/sites/default/files/logo.png.pagespeed.ce.rQZd6s4jcx.png\" alt=\"Porto\"></center>\r\n<TABLE WIDTH=\"87%\" BORDER=0 CELLSPACING=1 CELLPADDING=1>\r\n\t<TR>\r\n\r\n\t\t\r\n\t\t<TD width=\"1044\">\r\n\t\t<h2><font color=\"#0000FF\" face=\"Book Antiqua\">Exam Results - Apr/May \r\n\t\t2025 (</font><font face=\"Book Antiqua\" style=\"font-size: 15pt\" color=\"#008000\">All</font><font color=\"#008000\" face=\"Book Antiqua\"><span style=\"font-size: 15pt\"> \r\n\t\tStudents</span></font><font color=\"#0000FF\" face=\"Book Antiqua\">) </font>\r\n\t\t\r\n\t\t</h2>  \r\n\t\t</TD>\r\n\t\t<TD>\r\n\t\t<p align=\"center\"><b><a href=\"logout_new5.asp\">LOGOUT</a></b></TD>\t\t\r\n\t</TR>\r\n</TABLE>\r\n\r\n\r\n\r\n\r\n\r\n\r\n<TABLE border=3 borderColorDark=#c9beb6 borderColorLight=#000000 cellPadding=1 \r\ncellSpacing=1 width=\"70%\" align=center>\r\n\r\n\t<tr>\r\n\t\t<td ALIGN=\"middle\"><b>REG. NO.</b></td>\r\n\t\t<td ALIGN=\"middle\"><b>NAME</b></td>\r\n\t\t<td ALIGN=\"middle\"><b>COURSE</b></td>\r\n\r\n\t</tr>\r\n\r\n\t<tr>\r\n\t\t<td ALIGN=\"middle\">PRK24AD1007</td>\r\n\t\t<td ALIGN=\"middle\">SANJITH KRISHNA B</td>\r\n\t\t<td ALIGN=\"middle\">M.Sc.&nbsp;&nbsp;Artificial Intelligence and Data Science</td>\r\n\t\r\n\t</tr>\r\n</TABLE>\r\n<CENTER>\r\n<TABLE>\r\n\t<TR>\r\n\t\t<TD>\r\n\t\t\t<table BORDER=1 ALIGN=center>\r\n\t\t\t\t<tr>\r\n\t\t\t\t\t<td WIDTH=150 ALIGN=middle>Credits Earned</td> \r\n\t\t\t\t\t<td WIDTH=150 ALIGN=middle>CGPA</td> \r\n\t\t\t\t</tr>\r\n\t\t\t\t<tr>\r\n\t\t\t\t\t\r\n\t\t\t\t\t<TD ALIGN=middle><FONT COLOR=\"blue\" ><b>47</b></FONT></TD>\r\n\t\t\t\t\t\r\n\t\t\t\t\t<TD ALIGN=middle><FONT COLOR=\"blue\"><b>6.74</b></FONT></TD>\r\n\t\t\t\t\t\r\n\r\n\t\t\t\t</tr>\r\n\t\t\t</table>\r\n\t\t<TD WIDTH=10></TD>\r\n\t\t<TD>\r\n\t\t\t<table BORDER=1 ALIGN=center>\r\n\t\t\t\t<TR>\r\n\t\t\t\t\t<td WIDTH=150 ALIGN=middle>SGPA</td>\r\n\t\t\t\t</TR>\r\n\t\t\t\t<tr>\r\n\t\t\t\t\t\r\n\t\t\t\t\t<TD ALIGN=middle><FONT COLOR=\"blue\"><b>6.54</b></FONT></TD>\r\n\t\t\t\t\t\r\n\t\t\t\t\t\r\n\t\t\t\t\t\r\n\t\t\t\t\t\r\n\t\t\t\t</tr>\r\n\t\t\t</table>\r\n\t\t</TD>\r\n\t\t<TD WIDTH=50></TD>\r\n\r\n\t\t<td>\r\n\t\t<table BORDER=1 ALIGN=center>\r\n\t\t\t\t<TR>\r\n\t\t\t\t\t<td WIDTH=150 ALIGN=middle>Non 
Academic Credits Earned</td>\r\n\t\t\t\t</TR>\r\n\t\t\t\t<tr>\r\n\t\t\t\t\t<TD ALIGN=middle><FONT COLOR=\"blue\"><b>1</b></FONT></TD>\r\n\t\t\t\t\t\r\n\t\t\t\t\t\r\n\t\t\t\t\t\r\n\t\t\t\t</tr>\r\n\t\t\t</table>\r\n\r\n\t\t\r\n\t\t\r\n\t\t</td>\r\n\t\t\r\n\t\t\r\n\t\t\r\n\t\t\r\n\t</TR>\r\n</TABLE>\r\n\t\r\n\t\r\n\t\r\n</CENTER>\r\n<small>\r\n\r\n\r\n<!--\r\n<CENTER><FONT size=3><STRONG>\r\n<FONT color=green >Congratulations ! &nbsp;&nbsp;&nbsp; You have cleared all papers !</FONT>  </FONT>\r\n</CENTER></STRONG>\r\n-->\r\n\r\n\r\n<center>\r\n\r\n<TABLE border=3 borderColorDark=#c9beb6 borderColorLight=#000000 cellPadding=1 \r\ncellSpacing=1 width=\"65%\">\r\n\r\n\r\n  
<tbody>\r\n\r\n\r\n\r\n\r\n\r\n<tr ALIGN=\"middle\"><td><b>SEM</b></td><td><b>CODE</b></td><td><b>SUBJECT</b></td>\r\n<td><b>GRADE</b></td><td><b><b>CREDIT</b></b></td></tr></B>\r\n\r\n<td align=\"middle\"><font color=\"black\">  2</font></td>\r\n\r\n<td align=\"middle\"><font color=\"black\">  23CA3034</font></td>\r\n\r\n<td> <font color=\"black\"> Data Warehousing and Data Mining Techniques</font></td>\r\n\r\n\r\n<td align=\"middle\"><font color=\"BLACK\">P</font></td>\r\n\r\n\r\n<td align=\"middle\"><font color=\"black\">3</font></td>\r\n\r\n\r\n\r\n<td ALIGN=\"middle\"><font color=\"black\" \r\n      style=\"COLOR: fuchsia\">&nbsp; </font></td>\r\n\r\n\r\n</tr>\r\n\r\n\r\n<td align=\"middle\"><font color=\"black\">  2</font></td>\r\n\r\n<td align=\"middle\"><font color=\"black\">  23CA3036</font></td>\r\n\r\n<td> <font color=\"black\"> Data Analytics and Visualization</font></td>\r\n\r\n\r\n<td align=\"middle\"><font color=\"BLACK\">B+</font></td>\r\n\r\n\r\n<td align=\"middle\"><font color=\"black\">3</font></td>\r\n\r\n\r\n\r\n<td ALIGN=\"middle\"><font color=\"black\" \r\n      style=\"COLOR: fuchsia\">&nbsp; </font></td>\r\n\r\n\r\n</tr>\r\n\r\n\r\n<td align=\"middle\"><font color=\"black\">  2</font></td>\r\n\r\n<td align=\"middle\"><font color=\"black\">  23CA3038</font></td>\r\n\r\n<td> <font color=\"black\"> Advanced Database Technologies</font></td>\r\n\r\n\r\n<td align=\"middle\"><font color=\"BLACK\">B</font></td>\r\n\r\n\r\n<td align=\"middle\"><font color=\"black\">3</font></td>\r\n\r\n\r\n\r\n<td ALIGN=\"middle\"><font color=\"black\" \r\n      style=\"COLOR: fuchsia\">&nbsp; </font></td>\r\n\r\n\r\n</tr>\r\n\r\n\r\n<td align=\"middle\"><font color=\"black\">  2</font></td>\r\n\r\n<td align=\"middle\"><font color=\"black\">  23CA3039</font></td>\r\n\r\n<td> <font color=\"black\"> Human Centered Computing</font></td>\r\n\r\n\r\n<td align=\"middle\"><font color=\"BLACK\">B+</font></td>\r\n\r\n\r\n<td align=\"middle\"><font color=\"black\">3</font></td>\r\n\r\n\r\n\r\n<td ALIGN=\"middle\"><font color=\"black\" \r\n      style=\"COLOR: fuchsia\">&nbsp; </font></td>\r\n\r\n\r\n</tr>\r\n\r\n\r\n<td align=\"middle\"><font color=\"black\">  2</font></td>\r\n\r\n<td align=\"middle\"><font color=\"black\">  23CA3040</font></td>\r\n\r\n<td> <font color=\"black\"> Modelling Techniques in Predictive Analytics</font></td>\r\n\r\n\r\n<td align=\"middle\"><font color=\"BLACK\">B+</font></td>\r\n\r\n\r\n<td align=\"middle\"><font color=\"black\">3</font></td>\r\n\r\n\r\n\r\n<td ALIGN=\"middle\"><font color=\"black\" \r\n      style=\"COLOR: fuchsia\">&nbsp; </font></td>\r\n\r\n\r\n</tr>\r\n\r\n\r\n<td align=\"middle\"><font color=\"black\">  2</font></td>\r\n\r\n<td align=\"middle\"><font color=\"black\">  23CA3042</font></td>\r\n\r\n<td> <font color=\"black\"> Internet of Things and Blockchain Technology</font></td>\r\n\r\n\r\n<td align=\"middle\"><font color=\"BLACK\">P</font></td>\r\n\r\n\r\n<td align=\"middle\"><font color=\"black\">3</font></td>\r\n\r\n\r\n\r\n<td ALIGN=\"middle\"><font color=\"black\" \r\n      style=\"COLOR: fuchsia\">&nbsp; </font></td>\r\n\r\n\r\n</tr>\r\n\r\n\r\n<td align=\"middle\"><font color=\"black\">  2</font></td>\r\n\r\n<td align=\"middle\"><font color=\"black\">  22MA3001</font></td>\r\n\r\n<td> <font color=\"black\"> Logical Reasoning and Soft Skills</font></td>\r\n\r\n\r\n<td align=\"middle\"><font color=\"BLACK\">Completed</font></td>\r\n\r\n\r\n<td align=\"middle\"><font color=\"black\">0</font></td>\r\n\r\n\r\n\r\n<td ALIGN=\"middle\"><font color=\"black\" \r\n      style=\"COLOR: fuchsia\">&nbsp; </font></td>\r\n\r\n\r\n</tr>\r\n\r\n\r\n<td align=\"middle\"><font color=\"black\">  2</font></td>\r\n\r\n<td align=\"middle\"><font color=\"black\">  23CA3035</font></td>\r\n\r\n<td> <font color=\"black\"> Machine Learning Lab</font></td>\r\n\r\n\r\n<td align=\"middle\"><font color=\"BLACK\">B+</font></td>\r\n\r\n\r\n<td align=\"middle\"><font color=\"black\">2</font></td>\r\n\r\n\r\n\r\n<td ALIGN=\"middle\"><font color=\"black\" \r\n      style=\"COLOR: fuchsia\">&nbsp; </font></td>\r\n\r\n\r\n</tr>\r\n\r\n\r\n<td align=\"middle\"><font color=\"black\">  2</font></td>\r\n\r\n<td align=\"middle\"><font color=\"black\">  23CA3037</font></td>\r\n\r\n<td> <font color=\"black\"> Data Analytics and Visualization Lab</font></td>\r\n\r\n\r\n<td align=\"middle\"><font color=\"BLACK\">O</font></td>\r\n\r\n\r\n<td align=\"middle\"><font color=\"black\">2</font></td>\r\n\r\n\r\n\r\n<td ALIGN=\"middle\"><font color=\"black\" \r\n      style=\"COLOR: fuchsia\">&nbsp; </font></td>\r\n\r\n\r\n</tr>\r\n\r\n\r\n<td align=\"middle\"><font color=\"black\">  2</font></td>\r\n\r\n<td align=\"middle\"><font color=\"black\">  23CA3041</font></td>\r\n\r\n<td> <font color=\"black\"> Predictive Analytics Lab</font></td>\r\n\r\n\r\n<td align=\"middle\"><font color=\"BLACK\">A+</font></td>\r\n\r\n\r\n<td align=\"middle\"><font color=\"black\">2</font></td>\r\n\r\n\r\n\r\n<td ALIGN=\"middle\"><font color=\"black\" \r\n      style=\"COLOR: fuchsia\">&nbsp; </font></td>\r\n\r\n\r\n</tr>\r\n\r\n \r\n\r\n<center>\r\n<p>\r\n\r\n\r\n\r\n\r\n\r\n<BR></SMALL></TBODY></TABLE>\r\n<SMALL> <FONT color=darkviolet> The information at this site is provided solely for the user's immediate information and \r\nwithout warranty of any kind.  <BR>For final and correct data please confirm with the documents signed by the C.O.E.\r\n</FONT></SMALL><BR>\r\n\r\n\r\n\r\n\r\n<p align=right>\r\n\r\n<p align=left>\r\n\r\n<h2 align=\"center\">\r\n<span style=\"font-size: 15pt\">\r\n<i>\r\n&nbsp;<font color=\"#008000\"><a target=\"_blank\" href=\"CBCSSUPP/home_new.asp\">Supplementary \r\n&amp;Photocopy &amp; Revaluation \r\n- <font color=\"#008000\">Registration Link</font></a></font></i><br>\r\n</span>\r\n<font face=\"Garamond\" style=\"font-size: 13pt; font-weight: 700; font-style:italic; background-color:#FFFF00\" color=\"#008000\"><br>\r\n\t\t</font>\r\n\t\t<span style=\"font-size: 15pt\"><br>\r\n&nbsp;</span></h2>\r\n\r\n\r\n\r\n\r\n\r\n\r\n\r\n\r\n\r\n\r\n\r\n</font></b>\r\n\r\n\r\n\r\n\r\n\r\n<hr><center>\r\n\r\n<CENTER>\r\n\r\n\r\n<FONT size=3><BR>You are visitor No.<FONT color=green \r\nface=\"Comic Sans MS\">&nbsp;&nbsp;9798 </FONT>  to this page since \r\n01.07.2025\r\n\r\n<P>\r\n</P></FONT></CENTER></CENTER></H2></H2></u></font>\r\n\r\n\r\n<u>\r\n</FONT>\r\n\t\r\n\r\n<script defer src=\"https://static.cloudflareinsights.com/beacon.min.js/vcd15cbe7772f49c399c6a5babf22c1241717689176015\" integrity=\"sha512-ZpsOmlRQV6y907TI0dKBHq9Md29nnaEIPlkf84rnaERnq6zvWvPUqr2ft8M1aS28oN72PdrCzSjY4U6VaAw1EQ==\" data-cf-beacon='{\"rayId\":\"96308f83e8848600\",\"version\":\"2025.7.0\",\"serverTiming\":{\"name\":{\"cfExtPri\":true,\"cfEdge\":true,\"cfOrigin\":true,\"cfL4\":true,\"cfSpeedBrain\":true,\"cfCacheStatus\":true}},\"token\":\"d649c55a914642f792eb32a3436ff8de\",\"b\":1}' crossorigin=\"anonymous\"></script>\n</body>\r\n</html>"
}
]'''
    
    main(json_data=multiple_students_json)